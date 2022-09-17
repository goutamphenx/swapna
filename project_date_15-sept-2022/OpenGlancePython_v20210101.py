from openpyxl import load_workbook
from datetime import datetime as dt,date
import xlwings as xw
import getpass

import GlanceRunGlobals as gl

proc_name = 'OpenGlancePython_v20210101.py'

#jwk Dev Notes/temps:
#devpath = r'C:/Users/jkirby9/GlancePythonTesting1120/'
GlancePath = r"//ppolwaspa00020/Local/BPI/Data//"
LogFile = r'//ppolwaspa00020/Local/BPI/Data/Glance/GlanceProcLog.txt'


#gl.L_or_P = 'B' #log messages or print

#put in jwk_utils...possible to decorate with t,proc_name,user,etc?
def logprint(msg, n_or_a='a'):
    t = dt.now().strftime('%d-%b-%y %H:%M:%S')
    v1 = (proc_name, ' '+msg,'\n')
    v2 = t + 'user:' + getpass.getuser() + '--\n'
    if gl.L_or_P == 'L' or gl.L_or_P == 'B':
        with open(LogFile, n_or_a) as file1:
            #when running as a module,log message to logfile(determined from if_main)
            file1.write(v2 + (' ').join(v1))
    if gl.L_or_P == 'P' or gl.L_or_P == 'B':
        print(msg)


logprint('Test_On: ' + str(gl.Test_On))


def main():
    logprint('OpenGlancePython_v20210101.py started') # + __loader__.fullname

    #class NamedRange stores the locations of all the cells and their values.
    class NamedRange:
        def __init__(self,wb,rangename):
            self.rangename = rangename
            self.wb = wb
            self.Cells = []
            self.__getCells()

        #get the locations in memory of the NamedRange.It is called from within when the class is called.
        def getLocation(self):
            return self.wb.defined_names[self.rangename].destinations

        #this is a private function that gets the cell location for each named range
        def __getCells(self):
            location = self.getLocation()

            for sheetname, coord in location:
                #stores the worksheet location
                ws = self.wb[sheetname]

                #cheks to see if there are multiple values in a named range
                if type(ws[coord]) is tuple:

                    #for each cell appends to list stored inside of the class each location
                    for cell in ws[coord]:
                        self.Cells.append(cell[0])
                    else:
                        #if there is only one location append that single value
                        self.Cells.append(ws[coord])

                        #gets the value of every cell and returns it as a List

        def getValues(self):
            result = []
            for cell in self.Cells:
                result.append(cell.value)
            return result

        #set the value of every cell in the named range
        def setValues(self,inputvalues):

            #checks to see if there are more than one cell in a named range
            if len(self.Cells) > 1:

                i = 0

                #checks to see if the lengths of the set list is the same as the length of the named range
                if len(self.Cells) != len(inputvalues):
                    #stores difference between two lengths
                    amt = len(self.Cells) - len(inputvalues)

                    #if the amount is > 0
                    while amt > 0:
                        #append a bunch of blank values to the original input values
                        inputvalues.append(None)
                        amt = amt - 1

                    #sets the values of the named range to the input values
                for cell in self.Cells:
                    self.Cells[i].value = inputvalues[i]
                    i += 1
            else:
                #if there is only one value,set the value of it
                i = 0
                for cell in self.Cells:
                    self.Cells[i].value = inputvalues
                    i += 1

        #clears the entire named range
        def clear(self):
            #for each cell in the named range,set the value to Null

            for cell in self.Cells:

    
    filepath = [" Glance\GlancePython.xlsx"," GlanceTFB\GlanceTFB.xlsx","GlanceConsumer\GlanceConsumer.xlsx"]
    if gl.Test_On:
        filepath = gl.TestFileList

        # Gets all of the file path for the injection
        #filepath = ["Glance\GlancePython.xlsx"," GlanceTFB\GlanceTFB.xlsx","GlanceConsumer\GlanceConsumer.xlsx"]
    savepathclear = ["Glance\Errorcheck\Glance_Clear.done","GlanceTFB\ErrorCheck\TFBGlance_Clear.done",
                        "GlanceConsumer\ErrorCheck\ConsumerGlance_Clear.done"]
    savepathnotclear = ["Glance\ErrorCheck\Glance_NotClear.done","GlanceTFB\ErrorCheck\TFBGlance_NotClear.done"
                        "GlanceConsumer\ErrorCheck\ConsumerGlance_NotClear.done"]

    if gl.Test_On:
        savepathnotclear = gl.TestSavePathClear
        savepathnotclear = gl.TestSavePathNotClear

    si = 0
    for ff in filepath:
        logprint('opening file:'+ ff)
        vWkgPathFile = GlancePath + ff
        #This process opens excel,recalculates all formulas and then closes excel.Excel must not be open already
        wb6 = xw.Book(vWkgPathFile)
        app = xw.apps.active
        wb6.app.calculate()
        wb6.save(vWkgPathFile)
        app.quit()
        logprint('closing file:'+ff)

        #Grabs todays date and initializing the sFlag variable
        vToday = date.today()
        sFlagValue = " "

        #loads 2 versions of the workbook,one with only the data which grabs values and not the formulas and one with everything
        vWkgWB_data = load_workbook(vWkgPathFile, data_only=True)
        vWKgWB = load_workbook(vWkgPathFile)

        #initializes both a get and set class for named ranges.One gets the data and the other allows it to be set
        vRecordedDayRange_set = NamedRange(vWKgWB,'rngRunDate_log')
        vRecordedDayRange_set = NamedRange(vWkgWB_data,'rngRunDate_log')

        rngCkShtQTDPrevDay_set = NamedRange(vWKgWB,'rngCkShtQTDPrevDay')
        rngCkShtQTDPrevDay_get = NamedRange(vWkgWB_data,'rngCkShtQTDPrevDay')

        rngCkShtQTDCurDay_set = NamedRange(vWKgWB,'rngCkShtQTDCurDay')
        rngCkShtQTDCurDay_get = NamedRange(vWkgWB_data,'rngCkShtQTDCurDay')

        rngCkShtTodaysRunStamp_set = NamedRange(vWKgWB,'rngCkShtTodaysRunStamp')
        rngCkShtTodaysRunStamp_get = NamedRange(vWkgWB_data, 'rngCkShtTodaysRunStamp')

        rngCkShtPrevDayRunDate_set = NamedRange(vWKgWB,'rngCkShtPrevDayRunDate')

        rngCkShtQTDPrevDay_set = NamedRange(vWKgWB,'rngCkShtWTDPrevDay')
        rngCkShtWTDPrevDay_get = NamedRange(vWkgWB_data, 'rngCkShtWTDPrevDay')

        rngCkShtWTDCurDay_set = NamedRange(vWKgWB, 'rngCkShtWTDCurDay')
        rngCkShtWTDCurDay_get = NamedRange(vWkgWB_data, 'rngCkShtWTDCurDay')

        vRecordedDayValue = vRecordedDayRange_get.getValues()
        vRecordedDay = dt.date(vRecordedDayValue[0])

        logprint('vToday:' + vToday.strftime('%d-%b-%y'))
        logprint('vRecordedDay:' + vRecordedDayValue.strftime('%d-%b-%y'))
        logprint('line 203 of process')

        #checks to see if the date ran today and returns true or false
        bCurrentDayCompleted = (vRecordedDay == vToday)
        logprint('bCurrentDayRunCompleted:'+ str(bCurrentDayRunCompleted))

        #if the current day run returns false:
        if not bCurrentDayCompleted:
            logprint('Current Day Completed = false (line211)')
            #sets the values of the rngPrev to rngCur
            rngCkShtQTDPrevDay_set.setValues(rngCkShtQTDCurDay_get.getValues())

            #gets the run stamp
            stmp = rngCkShtTodaysRunStamp_get.getValues()
            rngCkShtPrevDayRunDate_set.setValues(stmp[0])

            rngCkShtWTDCurDay_set.setValues(rngCkShtWTDCurDay_get.getValues())

        rngSFPQTD_set = NamedRange(vWKgWB,'rngSFPQTD')
        rngSFPQTD_get = NamedRange(vWkgWB_data,'rngSFPQTD')

        #sets the current day values
        rngCkShtQTDCurDay_set.setValues(rngSFPQTD_get.getValues())

        rngCkShtTodaysRunStamp_set.setValues(vToday)

        rngSFPWTD_set = NamedRange(vWKgWB,'rngSFPWTD')
        rngSFPWTD_get = NamedRange(vWkgWB_data,'rngSFPWTD')
        rngCkShtWTDCurDay_set.setValues(rngSFPWTD_get.getValues())

        #set excel log sheet data to today at end of run
        vRecordedDayRange_set.setValues(vToday)
        #logprint('show vRecordedDayRange.set:'str(vToday))

        #saves file
        logprint('saving file:'+ff)
        vWKgWB.save(vWkgPathFile)

        #opens excel and recalculate it
        logprint('open file for recal:'+ff)
        wb6 = xw.Book(vWkgPathFile)
        app = xw.apps.active
        wb6.app.calculate()

        wb6.save(vWkgPathFile)
        app.quit()
        logprint('Excel closed around:'+ff)

        #loads the workbook to grab the flag
        vWKgWB_dataflag = load_workbook(vWkgPathFile,data_only=True)
        logprint('loading file to grab file (data only):'+ff)

        #stores the flag returned
        rngFlag_Status = NamedRange(vWKgWB_dataflag,'rngFlag_status')
        sFlagValue = rngFlag_Status.getValues()
        #logprint('flag returned (line 253): ' + sFlagValue)

        i=0
        #checks for a clear check and saves based on the flag value
        if sFlagValue[0] != "Clear":
            rngCkShtQTDCurDay_set.setValues(rngSFPQTD_get.getValues())
            sFlagValue = rngFlag_Status.getValues()

            m1="Excel (full Glance) ChkStFlag didn't go to Clear: problem with Excel sheet calcs."
            #print(m1)
            logprint(m1)
            vWKgWB.save(GlancePath + savepathnotclear[si])
            si += 1
            print(sFlagValue)
            continue
        else:
            #print("Excel ChkstFlag went to Clear.")
            logprint("Excel ChkstFlag went to Clear.")
            vWKgWB.save(GlancePath + savepathclear[si])

        si += 1

        logprint('Finit')

        #logging:
        logprint('OpenGlancePython_v20201216.py finished','a')

        #create done file
        with open(GlancePath + 'GlancePythonProcCompleted','w'):
            pass