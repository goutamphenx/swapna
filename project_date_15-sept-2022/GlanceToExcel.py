# -*- coding : utf-8 -*-
"""
Created on
@author:
"""

from TM1py.Services import TM1Service
from TM1py.Utils import Utils
import configparser
from datetime import datetime as dt

import getpass
import CalDateFuncs as cdf
import OpenGlancePython_v20210101 as ogp
import openpyxl as pxl

from pandas import ExcelWriter
import GlanceRunGlobals as gl

proc_name = 'ExpGlanceMDXINlineToExcel_Subscribers.py'
GlancePath = r'\\ppolwaspa0020\Local\BPI\Data\\'
mainGlancePath = GlancePath + 'Glance\\'

#create done file showing process has completed:(this is not a logmsg)
with open(mainGlancePath + 'FullPython_Glance_ProcStarted','w'):
    pass


LogFile = mainGlancePath + 'GlanceProcLog.txt'

#gl.L_or_P = 'B'

#put in jwk_utils... possible to decorate with t,proc_name,user,etc?
def logprint(msg,n_or_a = 'a'):
    t = dt.now().strftime('%d-%b-%y %H:%M:%S')
    v1 = (proc_name,' '+msg,'\n')
    v2 = t + 'user:'+getpass.getuser()+'--\n'
    if gl.L_or_P == 'L' or gl.L_or_P == 'B':
        with open(LogFile, n_or_a) as file1:
            #when running as module, log message to logfile(determined from if_main)
            file1.write(v2+(' ').join(v1))
    if gl.L_or_P == 'P' or gl.L_or_P == 'B':
        print(msg)


logprint('Test_On:' + str(gl.Test_On),'w')

'''
t=dt.now().strftime('%d-%b-%y %H:%M:%S')
with open(LogFile,"w") as file1:
        file1.write(t+ 'user:'+getpass.getuser()+'--\n' +
    'ExpglanceMDXtoExcel_v20201216.py started' )# +__loader__.fullname)
'''
logprint('ExpglanceMDXtoExcel_v20201216.py started')
logprint('module name.'+__name__)

config = configparser.ConfigParser()
##RV## config.read(r'.\config.ini')
config.read('C://Scripts//Config.ini')
#"C://Scripts//Config.ini"
#config.read('D://BPI//Data//Scripts//config.ini')
#config.read(r'.\config_jwk.ini')

srvArg = 'ppolwaspa00003'

#outfile_csv = r'E:\BPI\DataPythonGlanceData.csv'
#outfile_excel = GlancePath + 'GlancePython_SubsBackup.xlsx'
#outfile_excel = GlancePath + 'GlancePython.xlsx'
filepath = ["Glance/GlancePython.xlsx","GlanceTFB/GlanceTFB.xlsx","GlanceConsumer/GlanceConsumer.xlsx"]
if gl.Test_On:
    filepath = gl.TestFileList

outAlt = r'D:\BPI\Data\PythonGlanceDataAlt.csv'
outAlt = mainGlance + '\\' + 'PythonGlanceDataAlt.csv'


vcube,vview,vbp = 'Subscribers','SPEGS',False

'''
fully-formed string1 will look like the following,which will be place into MDX string:
{[Period - Day].[CurrentWeek],[Period - Day].[2020-05-16],[Period - Day].[2020-05-15],[Period - Day].[2020-05-14],[Period - Day].[2020-05-13],[Period - Day].[2020-05-12],[Period - Day].[2020-05-11],[Period - Day].[2020-05-10],[Period - Day].[QTD],[Period - Day].[WTD]}
'''
string1='{[Period - Day].[CurrentWeek],'
for d in cdf.A_curWeekAllDays():
    string1 += '[Period - Day].[' + d + '],'

#string1 += '[Period-Day].[QTD],[Period-Day].[WTD]}'
#changing 4.2.21 --jwk -
string1 += '[Period - Day].[QTD],[Period-Day].[WTD].[WTD QTR]}'

GLStringTmp1 = '''
  Select NON EMPTY TM1SubsetToSet([Account Segment Code],"SegmentPythonExtract") *
    TM1SubsetToSet([ML Rate Plan],"PythonML_Rate_Plan") *
    TM1SubsetToSet([Version - Customers],"VersionPythonExtract")*
    TM1SubsetToSet([Activation channel],"PythonActivationChannel")*
    TM1SubsetToSet([Product Line Code],"PythonExtractProduct")*
    TM1SubsetToSet([Measure - Subs],"PythonExtract")*
    VDATESTRING on ROWS,
    NON EMPTY TM1SubsetToSet([Deact Type],"PythonDeactType") on COLUMNS
    FROM [Subscribers]
    
    )'''

#leaving out:TM1SubsetToSet([Period - Day],"PeriodPythonExtract")*
vMDX = GLStringTmp1.replace('VDATESTRING',string1)

logprint(vMDX)

#def dfOutput()
with TM1Service(**config[srvArg]) as tm1:
    #Get Data from P&L cube through MDX
    pnl_data = tm1.cubes.cells.execute_mdx(vMDX)

    #Build Pandas Dataframe from raw cell set data
    df = Utils.build_pandas_dataframe_from_cellset(pn1_data,multiindex=False)

    newCol = df['Account Segment Code'] + df['ML Rate Plan'] + df['Version - Customers'] + \
             df['Activation Channel'] +df['Product Line Code'] + \
             df['Measure -Subs'] + df['Period - Day']

    df.insert(0, 'Concat',newCol)

logprint('finished grabbing tm1 data to DF')

#df.head(3)
#tmplogname = r'\\ppolwaspa00020\Local\BPI\Data\Glance\DataFrameGrab.txt'
tmplogname = r'c:\scripts\DataFrameGrab.txt'

try:
    df.to_csv(tmplogname, header=None, index=None, sep=' ')
except:
    logprint('DataFrame to .csv failed')
input('pausing here ...hitting a key will restart')

#properly load the workbook

for wb in filepath:
    file=GlancePath + wb

    logprint('line 116,file:' + file)
    #loads the saved file workbook into memory and goes to the tab
    wkbk = pxl.load_workbook(file,read_only=False,data_only=False, keep_links=False)
    tgtSht = wkbk['PythonGlanceData']

    #Inside this context manager,handle everything related to writing new data to the file\
    #without overwriting existing data

    #zero out all fields in dest worksheet PythonGlanceData...don't delete because potential linked cells
    mr2 = tgtSht.max_row
    mc2 = tgtSht.max_column

    for i in range(2,mr2+1):
        for j in range(1,mc2+1):
            tgtSht.Cell(i,j).value = ''

    #write the new data to new table in file; leaves existing tabs intact
    #loops through each value in dataframe and copies the value into the excel sheet of the saved workbook
    for i in range(len(df)):
        for j in range(len(df.columns)):
            if i == 0:
                tgtSht.Cell(i+1,j+1).value = df.columns[j]
            tgtSht.cell(i+2,j+1).value = df.iloc[i,j]

            #save workbook
            wkbk.save(file)
            wkbk.close()

''' for logging,pre Darrin's code'''
t = dt.now().strftime('%d-%b-%y %H:%,:%S')
with open(LogFile,"a") as file1:
    file1.write('/n' + t + 'user: ' + getpass.getuser()+ '--\n' +
                'ExpglanceMDXtoExcel_Subscribers_v20201216.py line121 \n')

# runs Darren's code to manipulate Excel
ogp.main()

''' for logging,post Darrin's code '''
logprint('ExpglanceMDXtoExcel_Subscribers_v20201216.py line130, just past importOpenGlancePython \n')

# create done file showing process has completed:(this is not a logmsg)
with open(mainGlancePath + '\\' + 'FullPython_Glance_ProComplted' , 'w'):
    pass

if __name__ == '__main__':
    pass
else:
    gl.L_or_P = 'B'
    # dfOutput()